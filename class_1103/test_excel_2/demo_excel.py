# -*- coding: utf-8 -*-
# @Time    : 2020/5/11 19:07
# @Author  : guoyunfei.0603

from openpyxl import load_workbook
from API_AUTO.tools.read_config import ReadConfig

class Demo_Excel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

# 第一种方式： all
    # def get_data(self,mode = 'all'):
        '''
        mode 改成 button也可以，约定的俗称    all
        mode:控制是否执行所有的用例，默认值为 all 为all就执行所有的用例
        如果不等于all 就进入分支判断
        mode 只能输入all 列表这两种类型的参数

        '''
# 第二种方式： 不加all，引入配置文件 read.config

    def get_data(self):
        # 从配置文件读取mode值
        mode = ReadConfig().read_config('case.config','MODE','mode')

        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        test_data = []
        for item in range(2,sheet.max_row+1):
            sub_data = {}

            sub_data['case_id'] = sheet.cell(item,1).value
            sub_data['url'] = sheet.cell(item,2).value
            sub_data['data'] = sheet.cell(item,3).value
            sub_data['expected'] = sheet.cell(item,4).value
            sub_data['method'] = sheet.cell(item,5).value
            test_data.append(sub_data)
        # return test_data
      # 根据button 去判断
        if mode == 'all': # 执行所有用例
            final_data = test_data
        else:#  [1,2,3,4] case_id
            final_data = []
            for item in test_data: # 对test_data 进行遍历

                # 如果用第二种方式，从配置文件读取的mode值[1,2,3,4] 是一个字符串，需要eval()
                if item['case_id'] in eval(mode): #
                    final_data.append(item)
        return final_data # 返回获取到的数据

if __name__ == '__main__':
    test_data = Demo_Excel('Demo.xlsx','Sheet1').get_data()
    print(test_data)

