# -*- coding: utf-8 -*-
# @Time    : 2020/5/8 11:41
# @Author  : guoyunfei.0603

from openpyxl import load_workbook

wb = load_workbook('前程贷注册、登录、充值.xlsx')

sheet = wb['lemon']

sub_data = {}

sub_data['test_num'] = sheet.cell(2,1).value

print(sub_data)