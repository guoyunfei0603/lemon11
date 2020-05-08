# -*- coding: utf-8 -*-
# @Time    : 2020/5/5 19:57
# @Author  : guoyunfei.0603

# 注意：不要直接在pycharm里面创建.xlsx文件！！
# 要在桌面或者其他位置创建Excel文件再复制到pycharm里面来
from openpyxl import load_workbook
# 1. 打开Excel
wb = load_workbook('qa.xlsx')

# 2. 定位表单
sheet = wb['test']

# 3. 定位单元格 -- 行列值
res = sheet.cell(2,2).value
print("拿到的结果是:",res)

print('最大行：',sheet.max_row)
print('最大列：',sheet.max_column)

for i in range(2,5):
    print('姓名:',sheet.cell(i,1).value)
