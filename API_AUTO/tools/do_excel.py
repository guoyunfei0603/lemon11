# -*- coding: utf-8 -*-
# @Time    : 2020/5/5 22:37
# @Author  : guoyunfei.0603

from openpyxl import load_workbook

class Do_excel:
    def __init__(self,file_name,sheet_name):# excel名字和表单名
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        # 1. 打开Excel
        wb = load_workbook(self.file_name)
        # 2. 定位表单
        sheet = wb[self.sheet_name]

        test_data = []

        for item in range(2,sheet.max_row+1):
            sub_data = {}
            sub_data['method'] = sheet.cell(item,1).value
            sub_data['url'] = sheet.cell(item,2).value
            sub_data['data'] = sheet.cell(item,3).value
            sub_data['headers'] = sheet.cell(item,4).value
            sub_data['expected'] = sheet.cell(item,5).value

            test_data.append(sub_data)
        return test_data # 返回获取到的数据

if __name__ == '__main__':
    print(Do_excel('lemon.xlsx','test').get_data())